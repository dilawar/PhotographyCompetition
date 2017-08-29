import numpy as np
from openpyxl import load_workbook


def final_score(val):
    if val >= 80:
        return 5
    elif val >= 60:
        return 4
    elif val >= 40:
        return 3
    elif val >= 20:
        return 2
    else:
        return 1


def calculate_score(file_name):
    wb2 = load_workbook(file_name)
    ws = wb2['Sheet1']
    names = []
    name_and_score = []
    score = []
    for w in ws.iter_cols(0, 1):
        for c in w:
            names.append(c.value)

    for w in ws.iter_cols(2, ws.max_column):
        n1 = []
        for c in w:
            n1.append(c.value)
        score.append(n1)
    score = np.asarray(score)
    total = []
    for j in score:
        min_value = min(j[j > 0])
        max_value = max(j[j > 0])
        m = [(x - min_value) * 100 / (max_value - min_value) if (x > 0) else 0 for x in j]
        for i in range(len(j)):
            if j[i] != 0:
                j[i] = final_score(m[i])

        total.append(j)
    total = np.asarray(total)
    for i in range(len(names)):
        name_and_score.append([names[i], sum(total[:, i]), total[:, i]])

    name_and_score = sorted(name_and_score, key=lambda l: l[1], reverse=True)
    print(name_and_score)
